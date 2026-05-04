"""
Recursive audit of the raw DataRails dump.

Scans data/raw/datarails_unzipped/ (both archives extracted side by side).
Profiles every CSV/XLSX. Classifies by likely_dataset_type. Computes date
coverage from candidate date columns. Calls out GL candidates and builds
a combined coverage report.

Outputs:
  data/staged/datarails_raw_file_inventory.csv
  data/staged/datarails_raw_file_inventory.md
  data/staged/gl_candidate_inventory.csv
  data/staged/gl_coverage_report.md
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO = Path("/Users/arench/Desktop/development_ontology_engine")
ROOT = REPO / "data" / "raw" / "datarails_unzipped"
STAGED = REPO / "data" / "staged"
STAGED.mkdir(parents=True, exist_ok=True)

# Heuristics
GL_NAME_HINTS = ["gl", "general ledger", "vertical financials", "transactions",
                 "journal", "activity", "qbo", "qbd", "quickbooks"]
GL_COL_HINTS_DR = {"PostingDate", "AccountCode", "FunctionalAmount", "DebitCredit", "TransactionNumber"}
GL_COL_HINTS_VF = {"Posting Date", "Account", "Debit", "Credit", "Amount"}
GL_COL_HINTS_QB = {"Type", "Date", "Debit", "Credit", "Memo"}  # QuickBooks register

CLICKUP_COL_HINTS = {"top_level_parent_id", "status", "date_created"}

DATE_NAME_RX = re.compile(
    r"(date|dt|posted|posting|created|approval|approved|due|start|end|"
    r"closed|close|sold|walk|c_of_o|cofo|fiscal|period)",
    re.I,
)
AMOUNT_NAME_RX = re.compile(
    r"(amount|amt|balance|debit|credit|total|cost|price|revenue|expense|principal|qty|quantity)",
    re.I,
)
ENTITY_NAME_RX = re.compile(r"(entity|company|companycode|companyname|llc|legal|subsidiary)", re.I)
ACCOUNT_NAME_RX = re.compile(r"(account|coa)", re.I)
PROJECT_NAME_RX = re.compile(r"(project|subdivision|community|jobname|customer:job)", re.I)
PHASE_NAME_RX = re.compile(r"(phase|stage|jobphase)", re.I)
LOT_NAME_RX = re.compile(r"(lot|address)", re.I)


def cols_match_any(cols: list[str], hints: set) -> bool:
    cs = set(c.strip() for c in cols)
    return len(cs & hints) >= max(2, int(0.6 * len(hints)))


def classify(path: Path, cols: list[str]) -> str:
    name = path.name.lower()
    cols_l = [c.strip() for c in cols]
    cols_set = set(cols_l)

    # ClickUp first (specific marker columns)
    if name.startswith("api-upload") or CLICKUP_COL_HINTS.issubset({c.lower() for c in cols_l}):
        return "ClickUp"

    # Inventory
    if "inventory" in name or "closing report" in name:
        return "inventory"

    # Allocation files
    if "allocation" in name:
        return "allocation"

    # Specific COA / TB / BS
    if "chart of accounts" in name or "coa" in name:
        return "COA"
    if "trial balance" in name or re.search(r"\btb\b", name):
        return "trial_balance"
    if "balance sheet" in name or "balance_sheet" in name or "combined bs" in name:
        return "balance_sheet"

    # AR
    if name.startswith("ar_") or "accounts receivable" in name or "ar aging" in name:
        return "AR"

    # GL — name OR columns
    if any(h in name for h in GL_NAME_HINTS):
        return "GL"
    if cols_match_any(cols_l, GL_COL_HINTS_DR):
        return "GL"
    if cols_match_any(cols_l, GL_COL_HINTS_VF):
        return "GL"
    # QB register-style: Type+Date+Debit+Credit must all be present
    if {"Type", "Date", "Debit", "Credit"}.issubset(cols_set):
        return "GL"

    # Collateral / underwriting / status reports
    if "collateral" in name:
        return "collateral"
    if "underwriting" in name:
        return "underwriting"
    if "status" in name:
        return "status_report"
    if "lot data" in name:
        return "lot_master"
    if "offer" in name:
        return "offer_master"
    if "cost to complete" in name:
        return "cost_to_complete"
    if "ia breakdown" in name:
        return "interest_alloc"
    if "rba" in name or "tnw" in name:
        return "summary"
    if "priorcr" in name:
        return "prior_cr"

    return "unknown"


def detect_columns(cols: list[str], rx: re.Pattern) -> list[str]:
    return [c for c in cols if rx.search(c)]


def best_header_for_xlsx(sheet_df_no_header: pd.DataFrame) -> int:
    """If row 0 looks like a title (mostly empty / one cell), find the first row
    where >=3 cells are non-empty AND look like headers."""
    for i in range(min(8, len(sheet_df_no_header))):
        row = sheet_df_no_header.iloc[i].astype(str).str.strip()
        nonempty = (row != "").sum()
        if nonempty >= 3:
            # Treat as header row if values are short-ish strings (not big numbers/dates)
            return i
    return 0


def read_csv_smart(path: Path, max_chunks: int = 100):
    """Header sniff then chunked iterator."""
    # Try header=0 first
    head = pd.read_csv(path, nrows=3, dtype=str, keep_default_na=False,
                       encoding="utf-8-sig", on_bad_lines="skip", encoding_errors="replace")
    cols = [str(c) for c in head.columns]
    # If a lot of columns are blank (title rows), try header=N up to 5
    if cols.count("") > len(cols) // 2 or any("Unnamed" in c for c in cols):
        for h in range(1, 6):
            try:
                head2 = pd.read_csv(path, nrows=3, header=h, dtype=str, keep_default_na=False,
                                    encoding="utf-8-sig", on_bad_lines="skip", encoding_errors="replace")
                cols2 = [str(c) for c in head2.columns]
                if cols2.count("") < len(cols2) // 3 and not any("Unnamed" in c for c in cols2[:5]):
                    return cols2, h
            except Exception:
                continue
    return cols, 0


def read_xlsx_smart(path: Path, sheet_name: str):
    """Try header=0; if it looks like a title row, scan."""
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
    h = best_header_for_xlsx(raw)
    df = pd.read_excel(path, sheet_name=sheet_name, header=h, dtype=str)
    return [str(c) for c in df.columns], h, df


def profile_dates(df: pd.DataFrame, candidate_date_cols: list[str]) -> tuple[str | None, str | None, dict]:
    """Return (min_iso, max_iso, per_col_stats) using best-effort parsing."""
    best_min, best_max = None, None
    per_col = {}
    for c in candidate_date_cols:
        if c not in df.columns:
            continue
        s = pd.to_datetime(df[c], errors="coerce")
        if s.notna().sum() == 0:
            per_col[c] = {"non_null": 0, "min": None, "max": None}
            continue
        mn, mx = s.min(), s.max()
        per_col[c] = {
            "non_null": int(s.notna().sum()),
            "min": mn.date().isoformat() if pd.notna(mn) else None,
            "max": mx.date().isoformat() if pd.notna(mx) else None,
        }
        if best_min is None or (mn and mn < pd.Timestamp(best_min)):
            best_min = mn.date().isoformat()
        if best_max is None or (mx and mx > pd.Timestamp(best_max)):
            best_max = mx.date().isoformat()
    return best_min, best_max, per_col


def profile_csv(path: Path) -> dict:
    out = {
        "rows": None, "cols": None, "columns": [], "header_row": 0, "sheet_name": None,
        "candidate_date": [], "candidate_amount": [], "candidate_entity": [],
        "candidate_account": [], "candidate_project": [], "candidate_phase": [],
        "candidate_lot": [], "min_date": None, "max_date": None, "per_date_col": {},
        "errors": [],
    }
    try:
        cols, header = read_csv_smart(path)
        out["columns"] = cols
        out["cols"] = len(cols)
        out["header_row"] = header

        out["candidate_date"] = detect_columns(cols, DATE_NAME_RX)
        out["candidate_amount"] = detect_columns(cols, AMOUNT_NAME_RX)
        out["candidate_entity"] = detect_columns(cols, ENTITY_NAME_RX)
        out["candidate_account"] = detect_columns(cols, ACCOUNT_NAME_RX)
        out["candidate_project"] = detect_columns(cols, PROJECT_NAME_RX)
        out["candidate_phase"] = detect_columns(cols, PHASE_NAME_RX)
        out["candidate_lot"] = detect_columns(cols, LOT_NAME_RX)

        # Stream for row count and date min/max
        rows_total = 0
        date_min, date_max = None, None
        per_col_stats = defaultdict(lambda: {"non_null": 0, "min": None, "max": None})

        usecols = out["candidate_date"]
        for chunk in pd.read_csv(
            path, header=header, chunksize=200_000, dtype=str, keep_default_na=False,
            usecols=usecols if usecols else None,
            on_bad_lines="skip", encoding="utf-8-sig", encoding_errors="replace",
        ):
            rows_total += len(chunk)
            for c in usecols:
                if c not in chunk.columns:
                    continue
                s = pd.to_datetime(chunk[c], errors="coerce")
                nn = int(s.notna().sum())
                if nn == 0:
                    continue
                mn, mx = s.min(), s.max()
                per_col_stats[c]["non_null"] += nn
                if per_col_stats[c]["min"] is None or mn.date().isoformat() < per_col_stats[c]["min"]:
                    per_col_stats[c]["min"] = mn.date().isoformat()
                if per_col_stats[c]["max"] is None or mx.date().isoformat() > per_col_stats[c]["max"]:
                    per_col_stats[c]["max"] = mx.date().isoformat()

        # If no date cols found, just count rows quickly
        if not usecols:
            rows_total = sum(1 for _ in open(path, "rb")) - 1 - header
            rows_total = max(rows_total, 0)

        out["rows"] = rows_total
        out["per_date_col"] = dict(per_col_stats)
        for c, st in per_col_stats.items():
            if st["min"] and (date_min is None or st["min"] < date_min):
                date_min = st["min"]
            if st["max"] and (date_max is None or st["max"] > date_max):
                date_max = st["max"]
        out["min_date"] = date_min
        out["max_date"] = date_max
    except Exception as e:
        out["errors"].append(f"{type(e).__name__}: {e}")
    return out


def profile_xlsx(path: Path) -> dict:
    out = {
        "rows": None, "cols": None, "columns": [], "header_row": 0, "sheet_name": None,
        "candidate_date": [], "candidate_amount": [], "candidate_entity": [],
        "candidate_account": [], "candidate_project": [], "candidate_phase": [],
        "candidate_lot": [], "min_date": None, "max_date": None, "per_date_col": {},
        "errors": [], "all_sheets": [],
    }
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        out["all_sheets"] = list(wb.sheetnames)
        wb.close()

        # Pick first non-empty sheet (largest by row count)
        best_sheet = out["all_sheets"][0]
        best_size = -1
        for sn in out["all_sheets"]:
            try:
                df_test = pd.read_excel(path, sheet_name=sn, header=None, dtype=str, nrows=200)
                # Count non-empty cells in first 200 rows as a proxy for "this sheet has data"
                n = int((df_test.astype(str) != "").sum().sum())
                if n > best_size:
                    best_size = n
                    best_sheet = sn
            except Exception:
                continue
        out["sheet_name"] = best_sheet

        cols, header, df = read_xlsx_smart(path, best_sheet)
        out["columns"] = cols
        out["cols"] = len(cols)
        out["rows"] = int(len(df))
        out["header_row"] = header

        out["candidate_date"] = detect_columns(cols, DATE_NAME_RX)
        out["candidate_amount"] = detect_columns(cols, AMOUNT_NAME_RX)
        out["candidate_entity"] = detect_columns(cols, ENTITY_NAME_RX)
        out["candidate_account"] = detect_columns(cols, ACCOUNT_NAME_RX)
        out["candidate_project"] = detect_columns(cols, PROJECT_NAME_RX)
        out["candidate_phase"] = detect_columns(cols, PHASE_NAME_RX)
        out["candidate_lot"] = detect_columns(cols, LOT_NAME_RX)

        mn, mx, per_col = profile_dates(df, out["candidate_date"])
        out["min_date"] = mn
        out["max_date"] = mx
        out["per_date_col"] = per_col
    except Exception as e:
        out["errors"].append(f"{type(e).__name__}: {e}")
    return out


def main():
    files = sorted([p for p in ROOT.rglob("*") if p.is_file() and not p.name.startswith(".")])
    print(f"Found {len(files)} files under {ROOT.relative_to(REPO)}")

    rows = []
    for i, p in enumerate(files, 1):
        ext = p.suffix.lower().lstrip(".")
        size = p.stat().st_size
        if ext == "csv":
            prof = profile_csv(p)
        elif ext in ("xlsx", "xlsm"):
            prof = profile_xlsx(p)
        else:
            prof = {"errors": [f"unsupported ext: {ext}"], "rows": None, "cols": None,
                    "columns": [], "candidate_date": [], "candidate_amount": [],
                    "candidate_entity": [], "candidate_account": [], "candidate_project": [],
                    "candidate_phase": [], "candidate_lot": [], "min_date": None,
                    "max_date": None, "per_date_col": {}, "header_row": 0, "sheet_name": None}

        likely = classify(p, prof.get("columns") or [])

        rows.append({
            "file_path": str(p.relative_to(REPO)),
            "file_name": p.name,
            "extension": ext,
            "size_bytes": size,
            "sheet_name": prof.get("sheet_name") or "",
            "header_row": prof.get("header_row", 0),
            "row_count": prof.get("rows"),
            "column_count": prof.get("cols"),
            "columns": "|".join(prof.get("columns") or []),
            "likely_dataset_type": likely,
            "min_date": prof.get("min_date") or "",
            "max_date": prof.get("max_date") or "",
            "candidate_date_columns": "|".join(prof.get("candidate_date") or []),
            "candidate_entity_columns": "|".join(prof.get("candidate_entity") or []),
            "candidate_amount_columns": "|".join(prof.get("candidate_amount") or []),
            "candidate_account_columns": "|".join(prof.get("candidate_account") or []),
            "candidate_project_columns": "|".join(prof.get("candidate_project") or []),
            "candidate_phase_columns": "|".join(prof.get("candidate_phase") or []),
            "candidate_lot_columns": "|".join(prof.get("candidate_lot") or []),
            "errors": "; ".join(prof.get("errors") or []),
            "_per_date_col": json.dumps(prof.get("per_date_col") or {}),
        })

    # Write inventory CSV
    inv_csv = STAGED / "datarails_raw_file_inventory.csv"
    with open(inv_csv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    # Save raw profiles for downstream use
    (STAGED / "datarails_raw_profiles.json").write_text(json.dumps(rows, indent=2, default=str))

    print(f"Wrote {inv_csv.relative_to(REPO)}")
    print(json.dumps({
        "n_files": len(rows),
        "by_type": {k: sum(1 for r in rows if r["likely_dataset_type"] == k)
                    for k in sorted({r["likely_dataset_type"] for r in rows})},
    }, indent=2))


if __name__ == "__main__":
    main()
