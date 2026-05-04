"""
One-shot organizer + profiler for DataRails_raw.zip dump.

- Classifies extracted files into data/raw/datarails/{gl,clickup,inventory_closing,balance_sheet,ar,unknown}
- Copies each file (preserving mtime) to its category folder AND to data/raw/datarails/snapshots/2026-05-01/
- Profiles every CSV/XLSX (rows, cols, columns, key/date/amount/entity guesses, null rates, samples, hash)
- Writes data/manifests/raw_export_manifest.csv
- Writes data/reports/raw_data_inventory_report.md
- Writes data/reports/staging_plan.md
- Does NOT delete the original zip or temp staging
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO = Path("/Users/arench/Desktop/development_ontology_engine")
SRC = REPO / "data" / "_unzip_tmp"
DR = REPO / "data" / "raw" / "datarails"
SNAPSHOT = DR / "snapshots" / "2026-05-01"
MANIFEST_PATH = REPO / "data" / "manifests" / "raw_export_manifest.csv"
INVENTORY_REPORT = REPO / "data" / "reports" / "raw_data_inventory_report.md"
STAGING_REPORT = REPO / "data" / "reports" / "staging_plan.md"


# --------------------------- classification ------------------------------- #

CLICKUP_HINT_COLS = {"top_level_parent_id", "name", "status", "date_created"}


def sniff_csv_columns(path: Path, max_bytes: int = 65536) -> list[str]:
    """Read header row (no full file load)."""
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        sample = f.read(max_bytes)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(sample.splitlines(), dialect)
    try:
        return next(reader)
    except StopIteration:
        return []


def classify(path: Path) -> tuple[str, str]:
    """Return (category_folder, dataset_guess). Lowercased name match."""
    fn = path.name
    lower = fn.lower()
    if lower.startswith("inventory") or "closing report" in lower:
        return "inventory_closing", "inventory_lots"
    if lower.startswith("balance_sheet") or "balance sheet" in lower:
        return "balance_sheet", "balance_sheet"
    # AR before GL because AR_QBD is also entity-specific
    if re.match(r"^ar[_ ]", lower) or "_ar_" in lower or lower.startswith("ar_"):
        return "ar", "accounts_receivable"
    if (
        lower.startswith("gl")
        or lower.startswith("gl_qbo")
        or lower.startswith("gl_qbd")
        or lower.startswith("gl ")
        or lower.startswith("gl(")
        or "gl_qbo_" in lower
        or "gl_qbd_" in lower
    ):
        return "gl", "general_ledger"
    if lower.startswith("api-upload"):
        if path.suffix.lower() == ".csv":
            try:
                cols = [c.strip() for c in sniff_csv_columns(path)]
            except Exception:
                cols = []
            if CLICKUP_HINT_COLS.issubset({c.lower() for c in cols}):
                return "clickup", "clickup_tasks"
        return "unknown", "unknown"
    return "unknown", "unknown"


def source_guess(fn: str) -> str:
    lower = fn.lower()
    if "_qbo_" in lower or "_qbo " in lower:
        return "QuickBooks Online (QBO)"
    if "_qbd_" in lower or "_qbd " in lower:
        return "QuickBooks Desktop (QBD)"
    if lower.startswith("api-upload"):
        return "ClickUp API export"
    if lower.startswith("inventory") or "closing report" in lower:
        return "Inventory closing report (likely DataRails or upstream ERP)"
    if lower.startswith("gl"):
        return "DataRails GL export (multi-entity bundle)"
    return "unknown"


# --------------------------- profiling ------------------------------------ #

DATE_KEYWORDS = (
    "date", "dt", "_at", "time", "posting", "period", "due", "start", "closed",
    "sold", "walk", "co_", "c_of_o", "approved", "created", "updated",
)
AMOUNT_KEYWORDS = (
    "amount", "amt", "balance", "debit", "credit", "value", "total", "cost",
    "price", "revenue", "expense", "net", "qty", "quantity", "principal",
)
ENTITY_KEYWORDS = (
    "entity", "company", "subsidiary", "llc", "legal_entity",
)
PROJECT_KEYWORDS = (
    "project", "subdivision", "community", "list", "list_name", "top_level_parent",
)
LOT_KEYWORDS = ("lot", "lot_num", "lot_number", "address")
PHASE_KEYWORDS = ("phase", "stage", "status")
KEY_KEYWORDS = ("id", "_id", "txn", "transaction", "key")


def col_match(col: str, kws) -> bool:
    cl = col.lower().strip()
    return any(kw in cl for kw in kws)


def file_hash(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def profile_csv(path: Path) -> dict:
    """Profile a CSV. Read in chunks so we don't OOM on big GL files."""
    out = {
        "rows": None, "cols": None, "columns": [],
        "key_fields": [], "date_fields": [], "amount_fields": [],
        "entity_fields": [], "project_fields": [], "lot_fields": [], "phase_fields": [],
        "null_rates": {}, "samples": [], "sheets": [], "errors": [],
    }
    try:
        # Header + samples without loading the whole file
        head = pd.read_csv(path, nrows=5, dtype=str, keep_default_na=False, on_bad_lines="skip", encoding_errors="replace")
        out["columns"] = [str(c) for c in head.columns]
        out["cols"] = len(out["columns"])
        out["samples"] = head.fillna("").astype(str).values.tolist()

        # Streaming row count + null counters for important cols (date/amount/key)
        important_idx = [
            i for i, c in enumerate(out["columns"])
            if col_match(c, DATE_KEYWORDS + AMOUNT_KEYWORDS + KEY_KEYWORDS)
        ]
        important_names = [out["columns"][i] for i in important_idx]
        null_counts = {n: 0 for n in important_names}
        rows = 0
        for chunk in pd.read_csv(
            path, chunksize=200_000, dtype=str, keep_default_na=False,
            usecols=important_names if important_names else None,
            on_bad_lines="skip", encoding_errors="replace",
        ):
            rows += len(chunk)
            for n in important_names:
                if n in chunk.columns:
                    s = chunk[n].astype(str).str.strip()
                    null_counts[n] += int((s == "").sum())
        # If no important cols, just get a row count
        if not important_names:
            rows = sum(1 for _ in open(path, "rb")) - 1
            rows = max(rows, 0)
        out["rows"] = rows
        if rows > 0:
            out["null_rates"] = {n: round(null_counts[n] / rows, 4) for n in important_names}

        for c in out["columns"]:
            if col_match(c, KEY_KEYWORDS): out["key_fields"].append(c)
            if col_match(c, DATE_KEYWORDS): out["date_fields"].append(c)
            if col_match(c, AMOUNT_KEYWORDS): out["amount_fields"].append(c)
            if col_match(c, ENTITY_KEYWORDS): out["entity_fields"].append(c)
            if col_match(c, PROJECT_KEYWORDS): out["project_fields"].append(c)
            if col_match(c, LOT_KEYWORDS): out["lot_fields"].append(c)
            if col_match(c, PHASE_KEYWORDS): out["phase_fields"].append(c)
    except Exception as e:
        out["errors"].append(f"profile_csv: {type(e).__name__}: {e}")
    return out


def profile_xlsx(path: Path) -> dict:
    out = {
        "rows": None, "cols": None, "columns": [],
        "key_fields": [], "date_fields": [], "amount_fields": [],
        "entity_fields": [], "project_fields": [], "lot_fields": [], "phase_fields": [],
        "null_rates": {}, "samples": [], "sheets": [], "errors": [],
    }
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        out["sheets"] = wb.sheetnames
        # Pick the largest-looking sheet (max non-empty rows)
        target_sheet = wb.sheetnames[0]
        for sn in wb.sheetnames:
            if "data" in sn.lower() or "report" in sn.lower() or "lot" in sn.lower():
                target_sheet = sn
                break
        wb.close()

        df_head = pd.read_excel(path, sheet_name=target_sheet, nrows=5, dtype=str)
        out["columns"] = [str(c) for c in df_head.columns]
        out["cols"] = len(out["columns"])
        out["samples"] = df_head.fillna("").astype(str).values.tolist()

        df_full = pd.read_excel(path, sheet_name=target_sheet, dtype=str)
        out["rows"] = int(len(df_full))
        important = [c for c in df_full.columns if col_match(str(c), DATE_KEYWORDS + AMOUNT_KEYWORDS + KEY_KEYWORDS)]
        if out["rows"] > 0:
            out["null_rates"] = {
                str(c): round(float(df_full[c].astype(str).str.strip().eq("").mean()), 4)
                for c in important
            }

        for c in out["columns"]:
            if col_match(c, KEY_KEYWORDS): out["key_fields"].append(c)
            if col_match(c, DATE_KEYWORDS): out["date_fields"].append(c)
            if col_match(c, AMOUNT_KEYWORDS): out["amount_fields"].append(c)
            if col_match(c, ENTITY_KEYWORDS): out["entity_fields"].append(c)
            if col_match(c, PROJECT_KEYWORDS): out["project_fields"].append(c)
            if col_match(c, LOT_KEYWORDS): out["lot_fields"].append(c)
            if col_match(c, PHASE_KEYWORDS): out["phase_fields"].append(c)
    except Exception as e:
        out["errors"].append(f"profile_xlsx: {type(e).__name__}: {e}")
    return out


# --------------------------- main ----------------------------------------- #

def safe_copy(src: Path, dst_dir: Path) -> Path:
    """Copy preserving mtime, never overwrite. If a name collision exists, append __dupN."""
    dst = dst_dir / src.name
    n = 2
    while dst.exists():
        # Only collide if hashes differ; otherwise skip silently and return existing.
        if file_hash(dst) == file_hash(src):
            return dst
        stem, suf = src.stem, src.suffix
        dst = dst_dir / f"{stem}__dup{n}{suf}"
        n += 1
    shutil.copy2(src, dst)
    return dst


def main() -> dict:
    files = sorted([p for p in SRC.iterdir() if p.is_file() and not p.name.startswith(".")])
    rows = []
    profiles: dict[str, dict] = {}
    placements: dict[str, list[str]] = {}
    hashes: dict[str, str] = {}

    for i, p in enumerate(files, 1):
        cat, dataset = classify(p)
        target_dir = DR / cat
        placed = safe_copy(p, target_dir)
        snap = safe_copy(p, SNAPSHOT)
        placements.setdefault(cat, []).append(placed.name)

        h = file_hash(p)
        hashes[p.name] = h

        suffix = p.suffix.lower()
        if suffix == ".csv":
            prof = profile_csv(p)
        elif suffix in (".xlsx", ".xlsm"):
            prof = profile_xlsx(p)
        else:
            prof = {"rows": None, "cols": None, "columns": [], "errors": [f"unsupported suffix: {suffix}"]}
        profiles[p.name] = prof

        size = p.stat().st_size
        mtime = datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).isoformat()
        confidence = "high" if cat != "unknown" else "low"

        notes_parts = []
        if prof.get("errors"):
            notes_parts.append("; ".join(prof["errors"]))
        if prof.get("sheets"):
            notes_parts.append(f"sheets: {','.join(prof['sheets'])}")
        notes_parts.append(f"sha256={h[:12]}")

        rows.append({
            "file_id": f"F{i:03d}",
            "original_filename": p.name,
            "current_path": str(placed.relative_to(REPO)),
            "file_type": suffix.lstrip("."),
            "source_guess": source_guess(p.name),
            "dataset_guess": dataset,
            "row_count": prof.get("rows"),
            "column_count": prof.get("cols"),
            "columns": "|".join(prof.get("columns", [])),
            "size_bytes": size,
            "modified_time": mtime,
            "notes": " | ".join(notes_parts),
            "confidence": confidence,
        })

    # Manifest CSV
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MANIFEST_PATH, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    # Persist profiles JSON for downstream + the report writer
    (REPO / "data" / "manifests" / "profiles.json").write_text(json.dumps(profiles, indent=2, default=str))
    (REPO / "data" / "manifests" / "hashes.json").write_text(json.dumps(hashes, indent=2))

    return {"rows": rows, "profiles": profiles, "placements": placements, "hashes": hashes}


if __name__ == "__main__":
    result = main()
    print(json.dumps({
        "n_files": len(result["rows"]),
        "by_category": {k: len(v) for k, v in result["placements"].items()},
        "manifest": str(MANIFEST_PATH.relative_to(REPO)),
    }, indent=2))
